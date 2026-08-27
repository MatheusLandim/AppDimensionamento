import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open('/tmp/fbv.json'))
amb = d['ambientes']; ues = d['ues']; pj = d['projeto']; c = d['cond']
wb = openpyxl.Workbook()

# estilos
AZUL = Font(name='Arial', color='0000FF', size=10)          # inputs
PRETO = Font(name='Arial', color='000000', size=10)         # fórmulas
BOLD = Font(name='Arial', bold=True, size=10)
TIT = Font(name='Arial', bold=True, size=14)
SUB = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HDRFILL = PatternFill('solid', fgColor='08404E')
YEL = PatternFill('solid', fgColor='FFF2CC')
ICE = PatternFill('solid', fgColor='D6E9EF')
thin = Side(style='thin', color='C9D3D8')
BOX = Border(left=thin,right=thin,top=thin,bottom=thin)
CEN = Alignment(horizontal='center', vertical='center')
def hdr(ws, row, cols):
    for i,t in enumerate(cols,1):
        cc=ws.cell(row=row,column=i,value=t); cc.font=SUB; cc.fill=HDRFILL; cc.alignment=CEN; cc.border=BOX

# ---------- CAPA ----------
cap = wb.active; cap.title='Capa'
cap['B2']='MEMORIAL DE CÁLCULO — CARGA TÉRMICA'; cap['B2'].font=TIT
cap['B3']='NBR 16401-1 · Método de carga por componentes'; cap['B3'].font=Font(name='Arial',size=10,italic=True,color='4A5D68')
info=[('Projeto',pj['nome']),('Código',pj['codigo']),('Local',pj['local']),('Cliente',pj['cliente']),('Revisão',pj['rev'])]
r=5
for k,v in info:
    cap.cell(r,2,k).font=BOLD; cap.cell(r,3,v).font=PRETO; r+=1
r+=1
cap.cell(r,2,'RESUMO').font=BOLD; r+=1
cap.cell(r,2,'Carga total do projeto').font=PRETO; cap.cell(r,3,'=Ambientes!C16').font=BOLD; cap.cell(r,3).number_format='#,##0" BTU/h"'; r+=1
cap.cell(r,2,'Carga total (TR)').font=PRETO; cap.cell(r,3,'=Ambientes!C16/12000').font=BOLD; cap.cell(r,3).number_format='#,##0.0" TR"'; r+=1
cap.cell(r,2,'Nº de ambientes').font=PRETO; cap.cell(r,3,len(amb)).font=PRETO; r+=1
cap.cell(r,2,'Condensadoras').font=PRETO; cap.cell(r,3,len(ues)).font=PRETO
cap.column_dimensions['B'].width=26; cap.column_dimensions['C'].width=28

# ---------- CONDIÇÕES ----------
co = wb.create_sheet('Condicoes')
co['A1']='CONDIÇÕES DE PROJETO'; co['A1'].font=TIT
rows=[('TBS externa (°C)',c['textBS'],'input'),('UR externa (%)',c['urExt'],'input'),
      ('TBS interna (°C)',c['tInt'],'input'),('UR interna (%)',c['urInt'],'input'),
      ('Altitude (m)',c['altitude'],'input'),
      ('Pressão atmosférica (kPa)','=101.325*(1-2.25577E-5*B6)^5.25588','f'),
      ('Fator de densidade','=B7/101.325','f'),
      ('Pvs externa (kPa)','=0.61094*EXP(17.625*B2/(B2+243.04))','f'),
      ('Pv externa (kPa)','=B4_0','skip'),
      ('Razão umidade ext (kg/kg)','=0.622*(B4pct)','skip'),
      ('Pvs interna (kPa)','=0.61094*EXP(17.625*B4/(B4+243.04))','f'),
      ('ΔT (ext-int) K','=B2-B4','f'),
      ('ΔW (ext-int) kg/kg','=MAX(0,B11-B14x)','skip')]
# escrevo manualmente p/ controlar referências
co['A2']='TBS externa (°C)'; co['B2']=c['textBS']
co['A3']='UR externa (%)'; co['B3']=c['urExt']
co['A4']='TBS interna (°C)'; co['B4']=c['tInt']
co['A5']='UR interna (%)'; co['B5']=c['urInt']
co['A6']='Altitude (m)'; co['B6']=c['altitude']
co['A7']='Pressão atmosférica (kPa)'; co['B7']='=101.325*(1-2.25577E-5*B6)^5.25588'
co['A8']='Fator de densidade'; co['B8']='=B7/101.325'
co['A9']='Pvs externa (kPa)'; co['B9']='=0.61094*EXP(17.625*B2/(B2+243.04))'
co['A10']='Pv externa (kPa)'; co['B10']='=B3/100*B9'
co['A11']='Razão umidade externa (kg/kg)'; co['B11']='=0.622*B10/(B7-B10)'
co['A12']='Pvs interna (kPa)'; co['B12']='=0.61094*EXP(17.625*B4/(B4+243.04))'
co['A13']='Pv interna (kPa)'; co['B13']='=B5/100*B12'
co['A14']='Razão umidade interna (kg/kg)'; co['B14']='=0.622*B13/(B7-B13)'
co['A15']='ΔT (ext - int) K'; co['B15']='=B2-B4'
co['A16']='ΔW (ext - int) kg/kg'; co['B16']='=MAX(0,B11-B14)'
for rr in range(2,7):
    co.cell(rr,2).font=AZUL; co.cell(rr,2).fill=YEL
for rr in range(7,17):
    co.cell(rr,2).font=PRETO
for rr in range(2,17):
    co.cell(rr,1).font=PRETO
co['B11'].number_format='0.00000'; co['B14'].number_format='0.00000'; co['B16'].number_format='0.00000'
co['B7'].number_format='0.000'; co['B8'].number_format='0.000'
co.column_dimensions['A'].width=30; co.column_dimensions['B'].width=16
co['D2']='Células amarelas = entradas editáveis.'; co['D2'].font=Font(name='Arial',italic=True,size=9,color='B26A12')

# ---------- COEFICIENTES ----------
cf = wb.create_sheet('Coeficientes')
cf['A1']='COEFICIENTES (valores-base editáveis — validar NBR/projeto)'; cf['A1'].font=TIT
consts=[('U parede (W/m²K)',2.5),('U vidro (W/m²K)',5.8),('U cobertura (W/m²K)',2.0),
        ('SHGC vidro',0.82),('ΔT cobertura sol-ar (K)',40),('c sensível ar ext',1.23),
        ('c latente ar ext',3010),('fator uso iluminação',0.85)]
for i,(k,v) in enumerate(consts,2):
    cf.cell(i,1,k).font=PRETO; cc=cf.cell(i,2,v); cc.font=AZUL; cc.fill=YEL
# orientação (A20:C23)
cf['A19']='Orientação'; cf['B19']='ΔTeq parede (K)'; cf['C19']='Radiação vidro (W/m²)'
for cc in ['A19','B19','C19']: cf[cc].font=BOLD
oris=[('N',14,130),('L',20,500),('S',10,90),('O',25,520)]
for i,(o,dt,rad) in enumerate(oris,20):
    cf.cell(i,1,o).font=PRETO; cf.cell(i,2,dt).font=AZUL; cf.cell(i,2).fill=YEL; cf.cell(i,3,rad).font=AZUL; cf.cell(i,3).fill=YEL
# pessoas (A26:C29)
cf['A25']='Atividade'; cf['B25']='Sensível (W)'; cf['C25']='Latente (W)'
for cc in ['A25','B25','C25']: cf[cc].font=BOLD
pes=[('repouso',65,35),('sentado_leve',70,45),('atividade_moderada',75,55),('exercicio',185,315)]
for i,(a,s,l) in enumerate(pes,26):
    cf.cell(i,1,a).font=PRETO; cf.cell(i,2,s).font=AZUL; cf.cell(i,2).fill=YEL; cf.cell(i,3,l).font=AZUL; cf.cell(i,3).fill=YEL
# categoria (A32:C34)
cf['A31']='Categoria'; cf['B31']='Ar ext L/s·pessoa'; cf['C31']='Ar ext L/s·m²'
for cc in ['A31','B31','C31']: cf[cc].font=BOLD
cat=[('residencia',2.5,0.3),('escritorio',2.5,0.3),('academia',10,0.3)]
for i,(a,pp,pa) in enumerate(cat,32):
    cf.cell(i,1,a).font=PRETO; cf.cell(i,2,pp).font=AZUL; cf.cell(i,2).fill=YEL; cf.cell(i,3,pa).font=AZUL; cf.cell(i,3).fill=YEL
cf.column_dimensions['A'].width=24; cf.column_dimensions['B'].width=18; cf.column_dimensions['C'].width=20

# ---------- CÁLCULO ----------
ca = wb.create_sheet('Calculo')
heads=['Tag','Ambiente','Área (m²)','Pessoas','Atividade','Categoria','Ilum (W/m²)','Equip (W)',
       'Parede (m²)','Orient.','Vidro (m²)','Cob.(0/1)',
       'Vazão AE (L/s)','Parede','Cobertura','Vidro cond.','Vidro rad.','Pessoas S','Ilum','Equip','Ar ext S',
       'SENSÍVEL (W)','Pessoas L','Ar ext L','LATENTE (W)','TOTAL (W)','TOTAL (BTU/h)','FCS','W/m²']
hdr(ca,1,heads)
for j,a in enumerate(amb):
    r=j+2
    vals=[a['tag'],a['nome'],a['area'],a['nPessoas'],a['atividade'],a['categoria'],a['iluminacaoWm2'],
          a['equipamentosW'],a['paredeExtArea'],a['orientacao'],a['vidroArea'],1 if a['coberturaExposta'] else 0]
    for i,v in enumerate(vals,1):
        cc=ca.cell(r,i,v); cc.font=AZUL; cc.border=BOX
        if i==3: cc.number_format='0.0'
    F=lambda col,f: (ca.cell(r,col,f))
    ca.cell(r,13,f'=INDEX(Coeficientes!$B$32:$B$34,MATCH(F{r},Coeficientes!$A$32:$A$34,0))*D{r}+INDEX(Coeficientes!$C$32:$C$34,MATCH(F{r},Coeficientes!$A$32:$A$34,0))*C{r}')
    ca.cell(r,14,f'=Coeficientes!$B$2*I{r}*INDEX(Coeficientes!$B$20:$B$23,MATCH(J{r},Coeficientes!$A$20:$A$23,0))')
    ca.cell(r,15,f'=IF(L{r}=1,Coeficientes!$B$4*C{r}*Coeficientes!$B$6,0)')
    ca.cell(r,16,f'=Coeficientes!$B$3*K{r}*Condicoes!$B$15')
    ca.cell(r,17,f'=K{r}*Coeficientes!$B$5*INDEX(Coeficientes!$C$20:$C$23,MATCH(J{r},Coeficientes!$A$20:$A$23,0))')
    ca.cell(r,18,f'=D{r}*INDEX(Coeficientes!$B$26:$B$29,MATCH(E{r},Coeficientes!$A$26:$A$29,0))')
    ca.cell(r,19,f'=G{r}*C{r}*Coeficientes!$B$9')
    ca.cell(r,20,f'=H{r}')
    ca.cell(r,21,f'=Coeficientes!$B$6_placeholder')  # substituído abaixo
    ca.cell(r,21,f'=Coeficientes!$B$6') # temp
    ca.cell(r,21,f'=Coeficientes!$B$6')
    ca.cell(r,21,f'=Coeficientes!$B$7*Condicoes!$B$8*M{r}*Condicoes!$B$15')  # ar ext sensível (cSens=Coef B7)
    ca.cell(r,22,f'=SUM(N{r}:U{r})')
    ca.cell(r,23,f'=D{r}*INDEX(Coeficientes!$C$26:$C$29,MATCH(E{r},Coeficientes!$A$26:$A$29,0))')
    ca.cell(r,24,f'=Coeficientes!$B$8*Condicoes!$B$8*M{r}*Condicoes!$B$16')  # ar ext latente (cLat=Coef B8)
    ca.cell(r,25,f'=W{r}+X{r}')
    ca.cell(r,26,f'=V{r}+Y{r}')
    ca.cell(r,27,f'=Z{r}*3.412142')
    ca.cell(r,28,f'=V{r}/Z{r}')
    ca.cell(r,29,f'=Z{r}/C{r}')
    for col in range(13,30):
        cc=ca.cell(r,col); cc.font=PRETO; cc.border=BOX
        if col in (27,): cc.number_format='#,##0'
        elif col==28: cc.number_format='0.00'
        elif col==29: cc.number_format='0'
        elif col==13: cc.number_format='0.0'
        else: cc.number_format='#,##0'
ca.freeze_panes='C2'
widths={'A':7,'B':18,'E':18,'F':14}
for k,w in widths.items(): ca.column_dimensions[k].width=w

# ---------- AMBIENTES (resumo) ----------
am = wb.create_sheet('Ambientes')
hdr(am,1,['Tag','Ambiente','Carga (BTU/h)','FCS','W/m²','UI selecionada','UE'])
for j,a in enumerate(amb):
    r=j+2
    am.cell(r,1,f'=Calculo!A{r}'); am.cell(r,2,f'=Calculo!B{r}')
    am.cell(r,3,f'=Calculo!AA{r}'); am.cell(r,3).number_format='#,##0'
    am.cell(r,4,f'=Calculo!AB{r}'); am.cell(r,4).number_format='0.00'
    am.cell(r,5,f'=Calculo!AC{r}'); am.cell(r,5).number_format='0'
    am.cell(r,6,a['ui']); am.cell(r,7,a['ue'])
    for col in range(1,8): am.cell(r,col).font=PRETO; am.cell(r,col).border=BOX
rtot=len(amb)+2
am.cell(rtot,2,'TOTAL').font=BOLD
am.cell(rtot,3,f'=SUM(C2:C{rtot-1})').font=BOLD; am.cell(rtot,3).number_format='#,##0'; am.cell(rtot,3).fill=ICE
am.column_dimensions['B'].width=18; am.column_dimensions['F'].width=15
# célula usada pela Capa: total em C16
if rtot!=16:
    am.cell(16,2,'(total)'); am.cell(16,3,f'=SUM(C2:C{len(amb)+1})'); am.cell(16,3).number_format='#,##0'

# ---------- EQUIPAMENTOS ----------
eq = wb.create_sheet('Equipamentos')
hdr(eq,1,['Tag','Ambiente','Carga (BTU/h)','UI','Cap. UI (BTU/h)','UE'])
for j,a in enumerate(amb):
    r=j+2
    eq.cell(r,1,a['tag']); eq.cell(r,2,a['nome']); eq.cell(r,3,f'=Ambientes!C{r}'); eq.cell(r,3).number_format='#,##0'
    eq.cell(r,4,a['ui']); eq.cell(r,5,a['uiCap']); eq.cell(r,5).number_format='#,##0'; eq.cell(r,6,a['ue'])
    for col in range(1,7): eq.cell(r,col).font=PRETO; eq.cell(r,col).border=BOX
base=len(amb)+3
eq.cell(base,1,'BALANÇO POR CONDENSADORA').font=BOLD
hdr(eq,base+1,['UE','Modelo','Nominal (BTU/h)','Soma UIs (BTU/h)','Taxa conexão','Qtd UIs'])
for k,u in enumerate(ues):
    r=base+2+k
    eq.cell(r,1,u['tag']); eq.cell(r,2,u['modelo']); eq.cell(r,3,u['nominal']); eq.cell(r,3).number_format='#,##0'
    eq.cell(r,4,f'=SUMIF($F$2:$F${len(amb)+1},A{r},$E$2:$E${len(amb)+1})'); eq.cell(r,4).number_format='#,##0'
    eq.cell(r,5,f'=D{r}/C{r}'); eq.cell(r,5).number_format='0.0%'
    eq.cell(r,6,u['qtd'])
    for col in range(1,7): eq.cell(r,col).font=PRETO; eq.cell(r,col).border=BOX
eq.column_dimensions['B'].width=16; eq.column_dimensions['C'].width=16; eq.column_dimensions['D'].width=16; eq.column_dimensions['E'].width=13

# ---------- MEMORIAL ----------
me = wb.create_sheet('Memorial')
me['A1']='MEMORIAL JUSTIFICATIVO'; me['A1'].font=TIT
paras=[
 f"O presente memorial apresenta o dimensionamento da carga térmica de refrigeração do projeto {pj['nome']} ({pj['codigo']}), {pj['local']}, cliente {pj['cliente']}, revisão {pj['rev']}.",
 f"O cálculo seguiu o método de carga por componentes conforme a NBR 16401-1, adotando condições externas de projeto de {c['textBS']}°C TBS / {c['urExt']}% UR e condições internas de conforto de {c['tInt']}°C / {c['urInt']}% UR (NBR 16401-2). A densidade do ar foi corrigida para a altitude de {c['altitude']} m, ajustando as parcelas de renovação de ar externo.",
 "As parcelas consideradas por ambiente incluem: condução e insolação pelas vedações opacas e envidraçamentos (com diferença de temperatura equivalente sol-ar por orientação), ganhos internos de pessoas, iluminação e equipamentos, e a carga de renovação de ar externo (sensível e latente) conforme NBR 16401-3.",
 f"A carga térmica total resultante do projeto é de {d['totalBtu']:,.0f} BTU/h ({d['totalTR']:.2f} TR), distribuída entre {len(amb)} ambientes e {len(ues)} unidades condensadoras VRF.".replace(',','.'),
 "A seleção das unidades internas adotou o modelo de menor capacidade que atende à carga de cada ambiente. O balanço por condensadora (aba Equipamentos) verifica a taxa de conexão em relação à capacidade nominal, garantindo operação dentro dos limites do fabricante.",
 "Os coeficientes de transmitância, fator solar, diferença sol-ar e radiação (aba Coeficientes) são valores-base e devem ser confirmados conforme a construção e as esquadrias efetivamente especificadas.",
]
r=3
for p in paras:
    cell=me.cell(r,1,p); cell.font=Font(name='Arial',size=10); cell.alignment=Alignment(wrap_text=True,vertical='top')
    me.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8); me.row_dimensions[r].height=46; r+=2
me.column_dimensions['A'].width=14
for col in 'BCDEFGH': me.column_dimensions[col].width=14

wb.save('/mnt/user-data/outputs/Memorial_FBV.xlsx')
print("salvo. abas:", wb.sheetnames)
